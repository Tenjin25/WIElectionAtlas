from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = ROOT / "data"
GENERIC_TITLES = {
    "WEC Canvass Reporting System",
    "Ward by Ward Report",
    "County by County Report",
}
WORKBOOK_CONFIG = {
    "Ward by Ward Report__April 2 2024 Spring Election_All Contests.xlsx": {
        "date": "20240402",
        "election_type": "general",
        "suffix": "ward",
    },
    "Ward by Ward Report_November 5 2024 General Election_Federal and State Contests.xlsx": {
        "date": "20241105",
        "election_type": "general",
        "suffix": "ward",
    },
    "Ward by Ward Report by Congressional District_November 5 2024 General Election_Federal and State Contests.xlsx": {
        "date": "20241105",
        "election_type": "general",
        "suffix": "ward-by-congressional-district",
    },
    "April 1 2025 Spring Election_Ward by Ward_State Level Contests and Referendum.xlsx": {
        "date": "20250401",
        "election_type": "general",
        "suffix": "ward",
    },
    "April 1 2025 Spring Election_County by County Report.xlsx": {
        "date": "20250401",
        "election_type": "general",
        "suffix": "county",
    },
    "Ward by Ward Report_Spring Election 2026_All State Contests.xlsx": {
        "date": "20260407",
        "election_type": "general",
        "suffix": "ward",
    },
    "County by County Report_Spring Election 2026_All State Contests.xlsx": {
        "date": "20260407",
        "election_type": "general",
        "suffix": "county",
    },
}


def cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def first_nonempty(values: Iterable[object]) -> str:
    for value in values:
        text = cell_text(value)
        if text:
            return text
    return ""


def titlecase_geo(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip()).title()


def clean_candidate(value: str) -> str:
    collapsed = re.sub(r"\s+", " ", value.strip())
    if collapsed.upper() == "SCATTERING":
        return "Scattering"
    return collapsed


def normalize_party(value: str, candidate: str) -> str:
    if candidate == "Scattering":
        return ""
    return value.strip()


def find_header_row(sheet) -> int:
    for row_idx in range(1, 26):
        values = [cell_text(cell) for cell in next(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))]
        if any(value == "Total Votes Cast" for value in values):
            return row_idx
    raise ValueError(f"Could not find header row in {sheet.title}")


def find_contest_name(sheet, header_row: int) -> str:
    for row_idx in range(header_row - 1, 0, -1):
        row = next(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        for value in row:
            text = cell_text(value)
            if text and text not in GENERIC_TITLES and "Election" not in text:
                return text
    return sheet.title


def normalize_office(contest_name: str) -> tuple[str, str]:
    contest = re.sub(r"\s+", " ", contest_name.strip())
    upper = contest.upper()

    matchers = [
        (r"^PRESIDENT OF THE UNITED STATES.*", "President"),
        (r"^UNITED STATES SENATOR$", "Senate"),
        (r"^REPRESENTATIVE IN CONGRESS DISTRICT (\d+).*", "House"),
        (r"^STATE ASSEMBLY DISTRICT (\d+).*", "State Assembly"),
        (r"^STATE SENATE DISTRICT (\d+).*", "State Senate"),
        (r"^COURT OF APPEALS JUDGE DISTRICT (\d+).*", "Court Of Appeals Judge"),
    ]

    for pattern, office in matchers:
        match = re.match(pattern, upper)
        if match:
            district = match.group(1) if match.groups() else ""
            return office, district

    if upper.startswith("JUSTICE OF THE SUPREME COURT"):
        return "Supreme Court", ""

    if upper.startswith("STATE SUPERINTENDENT OF PUBLIC INSTRUCTION"):
        return "State Superintendent Of Public Instruction", ""

    return contest.title(), ""


def parse_sheet(path: Path, sheet) -> list[dict[str, object]]:
    config = WORKBOOK_CONFIG[path.name]
    header_row = find_header_row(sheet)
    header_top = [cell_text(value) for value in next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    header_bottom = [cell_text(value) for value in next(sheet.iter_rows(min_row=header_row + 1, max_row=header_row + 1, values_only=True))]
    contest_name = find_contest_name(sheet, header_row)
    office, district = normalize_office(contest_name)

    rows: list[dict[str, object]] = []
    current_county = ""

    for values in sheet.iter_rows(min_row=header_row + 2, values_only=True):
        cleaned = [cell_text(value) for value in values]
        if not any(cleaned):
            continue

        total_votes = values[2] if len(values) > 2 else None
        if total_votes in (None, ""):
            continue

        raw_county = cleaned[0] if len(cleaned) > 0 else ""
        raw_ward = cleaned[1] if len(cleaned) > 1 else ""

        if config["suffix"].startswith("ward"):
            if raw_county:
                current_county = raw_county
            county = titlecase_geo(current_county)
            ward = titlecase_geo(raw_ward)
        else:
            county = titlecase_geo(raw_county)
            ward = ""

        for col_idx in range(3, len(values)):
            candidate_name = clean_candidate(header_bottom[col_idx]) if col_idx < len(header_bottom) else ""
            party = header_top[col_idx].strip() if col_idx < len(header_top) else ""
            votes = values[col_idx]

            if not candidate_name or votes in (None, ""):
                continue

            rows.append(
                {
                    "county": county,
                    "ward": ward,
                    "office": office,
                    "district": district,
                    "total.votes": total_votes,
                    "party": normalize_party(party, candidate_name),
                    "candidate": candidate_name,
                    "votes": votes,
                }
            )

    return rows


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    fieldnames = ["county", "ward", "office", "district", "total.votes", "party", "candidate", "votes"]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    wrote: list[tuple[Path, int]] = []

    for workbook_name, config in WORKBOOK_CONFIG.items():
        matches = list(DATA_ROOT.glob(f"*/{workbook_name}"))
        if not matches:
            raise FileNotFoundError(f"Workbook not found: {workbook_name}")

        workbook_path = matches[0]
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        rows: list[dict[str, object]] = []

        for sheet_name in workbook.sheetnames:
            if sheet_name == "Document map":
                continue
            rows.extend(parse_sheet(workbook_path, workbook[sheet_name]))

        output_name = f"{config['date']}__wi__{config['election_type']}__{config['suffix']}.csv"
        output_path = workbook_path.parent / output_name
        write_csv(output_path, rows)
        wrote.append((output_path, len(rows)))

    for output_path, row_count in wrote:
        print(f"Wrote {row_count} rows to {output_path}")


if __name__ == "__main__":
    main()
